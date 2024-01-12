# Nichoals Wharton
# KEEPA Scraping Program
# 12/17 updated version

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from multiprocessing import Process, Lock
import multiprocessing
import openpyxl
import pandas
import time
import sys

inputSheet = 'SEETSDataSheet12-22-17-02.xlsx'
noDataLimit = 15

def loop(lock, driver, i, noDataVal):
    
    try:
        #if data cannt be found for the item then ignore the exception and continue with the next entry
        lock.acquire()
        book = openpyxl.load_workbook(inputSheet)
        sheet = book.active

        #get the asin value for the entry
        asin = sheet.cell(row=i, column=3).value
        print(str(i - 1) + ": " + asin)
    except KeyboardInterrupt:
        print("Interrupted by user, saving workbook...")
        sys.exit(1)
    finally:
        book.save(inputSheet)
        lock.release()
        
    #open its corresponding keepa page
    driver.get("https://keepa.com/#!product/1-" + asin)

    #if the element is not loaded wait up to 50 seconds for it to complete
    driver.implicitly_wait(25)
    
    try:
        
        #wait until the data tab is clickable, then explicitly wait for 2 seconds, then select the element and click it
        WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, 'tabMore')))
        time.sleep(2)
        dataTab2 = driver.find_element(By.ID, 'tabMore')
        dataTab2.click()
        
        #wait until the offers tab is clickable, then explicitly wait for 2 seconds, then select the element and click it
        WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="tabHeadSub"]/li[2]')))
        time.sleep(2)
        offersTab2 = driver.find_element(By.XPATH, '//*[@id="tabHeadSub"]/li[2]')
        offersTab2.click()

        hOffers = driver.find_element(By.XPATH, '//*[@id="ft-includeHistoricalOffers"]')
        hOffers.click()

        avgSpan = driver.find_element(By.XPATH, '//*[@id="grid-offer"]/div/div[2]/div[1]/div[4]/div[2]/div/div/div[1]/span')
        avgStr = ''

        for ch in avgSpan.text:
            if ch.isdigit():
                avgStr += ch

        stockTotal = driver.find_element(By.XPATH, '//*[@id="grid-offer"]/div/div[2]/div[1]/div[4]/div[2]/div/div/div[4]/span')
        stStr = ''

        for ch in stockTotal.text:
            if ch.isdigit():
                stStr += ch

        #find the total sold span, get its value and then seperate the digits from the label
        sold = driver.find_element(By.XPATH, '//*[@id="grid-offer"]/div/div[2]/div[1]/div[4]/div[2]/div/div/div[5]/span')
        soldStr = ''

        for ch in sold.text:
            if ch.isdigit():
                soldStr += ch

        #find the total sold over the last 30 days span, get its value and then seperate the digits from the label
        sold30 = driver.find_element(By.XPATH, '//*[@id="grid-offer"]/div/div[2]/div[1]/div[4]/div[2]/div/div/div[6]/span')
        sold30Str = ''

        for ch in sold30.text:
            if ch.isdigit():
                sold30Str += ch

        try:
            #add the found values to the spreadsheet
            lock.acquire()
            book = openpyxl.load_workbook(inputSheet)
            sheet = book.active
            sheet.cell(row=i, column=26, value=avgStr) #average sales price
            sheet.cell(row=i, column=31, value=stStr) #stock total
            sheet.cell(row=i, column=30, value=soldStr) #total sales
            sheet.cell(row=i, column=39, value=sold30Str) #30 day sales
        except KeyboardInterrupt:
            print("Interrupted by user, saving workbook...")
            sys.exit(1)
        finally:
            book.save(inputSheet)
            lock.release()
        


    except:
        #runs if an exception was thrown attempting to recieve the entrys data
        print("no data")
        with noDataVal.get_lock():
            noDataVal.value += 1
            print("noDataVal: " + str(noDataVal.value))
        book.save(inputSheet)
            


def p1(lock, fileRows, offset, noDataVal):
    options = Options()
    options.add_argument("window-size=1920,1080")
    driver = webdriver.Chrome(options=options)
    driver.get("https://keepa.com/#!product/1-" + "B00005NUZN")
    time.sleep(5)
    loginButton = driver.find_element(By.XPATH, '//*[@id="panelUserRegisterLogin"]')
    loginButton.click()

    time.sleep(5)
    usernameInput = driver.find_element(By.XPATH, '//*[@id="username"]')
    usernameInput.send_keys("-!!!- ADD YOUR USERNAME HERE -!!!-")

    passwordInput = driver.find_element(By.XPATH, '//*[@id="password"]')
    passwordInput.send_keys("-!!!- ADD YOUR PASSWORD HERE -!!!-")

    submitLogin = driver.find_element(By.XPATH, '//*[@id="submitLogin"]')
    submitLogin.click()

    time.sleep(5)

    i = 2 + int(offset)
    #loop through each SEETS entry in the data sheet
    while i < fileRows:
        if noDataVal.value > noDataLimit:
            break

        #if (i % 3) == 0:
        loop(lock, driver, i, noDataVal)
        i += 1
    driver.quit()


if __name__ == '__main__':
    #determines the amount of SEETS entries to search for
    ds = pandas.read_excel(inputSheet)
    fileRows = ds.shape[0] + 2

    lock = Lock()

    print("fileRows: " + str(fileRows))

    offset = ''
    while offset.isdigit() == False:
        offset = input("Offset Amount: (0 if you want to start from the beginning) ")

    #adds the labels to the spreadsheet for the new attributes
    book = openpyxl.load_workbook(inputSheet)
    sheet = book.active

    book.save(inputSheet)

    noDataVal = multiprocessing.Value('i', 0)

    process1 = Process(target=p1, args=(lock, fileRows, offset, noDataVal))

    process1.start()

    process1.join()