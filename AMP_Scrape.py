# Nichoals Wharton
# AMP Scraping Program
# 12/17 updated version

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from multiprocessing import Process, Lock
import multiprocessing
import openpyxl
import pandas
import time
import sys

inputSheet = 'SEETSDataSheet12-22-17-02.xlsx'
noDataLimit = 15

def loop(lock, driver, i, noDataVal):
    
    try:
        #if data cannt be found for the item then ignore the exception and continue with the next entry
        lock.acquire()
        book = openpyxl.load_workbook(inputSheet)
        sheet = book.active

        #get the asin value for the entry
        asin = sheet.cell(row=i, column=3).value
        print(str(i - 1) + ": " + asin)
    except KeyboardInterrupt:
        print("Interrupted by user, saving workbook...")
        sys.exit(1)
    finally:
        book.save(inputSheet)
        lock.release()
        
    #open its corresponding keepa page
    driver.get("https://sas.selleramp.com/sas/lookup?SasLookup%5Bsearch_term%5D=" + asin)

    #if the element is not loaded wait up to 50 seconds for it to complete
    driver.implicitly_wait(25)
    
    try:
        lowestFBA = driver.find_element(By.XPATH, '//*[@id="keepa_csv_type_10"]/span')
        lowestFBAStr = lowestFBA.text

        lowestFBM = driver.find_element(By.XPATH, '//*[@id="keepa_csv_type_7"]/span')
        lowestFBMStr = lowestFBM.text

        maxCost = driver.find_element(By.XPATH, '//*[@id="saslookup-max-cost"]')
        maxCostStr = maxCost.text
        
        
        try:
            #add the found values to the spreadsheet
            lock.acquire()
            book = openpyxl.load_workbook(inputSheet)
            sheet = book.active
            sheet.cell(row=i, column=20, value=maxCostStr)
            sheet.cell(row=i, column=22, value=lowestFBAStr)
            sheet.cell(row=i, column=23, value=lowestFBMStr)
        except KeyboardInterrupt:
            print("Interrupted by user, saving workbook...")
            sys.exit(1)
        finally:
            book.save(inputSheet)
            lock.release()
        
    except:
        #runs if an exception was thrown attempting to recieve the entrys data
        print("no data")
        with noDataVal.get_lock():
            noDataVal.value += 1
            print("noDataVal: " + str(noDataVal.value))
        book.save(inputSheet)
            


def p1(lock, fileRows, offset, noDataVal):
    options = Options()
    options.add_argument("window-size=1920,1080")
    driver = webdriver.Chrome(options=options)
    driver.get("https://sas.selleramp.com/sas/lookup?SasLookup%5Bsearch_term%5D=" + "B00N58KQMQ")

    time.sleep(5)
    usernameInput = driver.find_element(By.XPATH, '//*[@id="loginform-email"]')
    usernameInput.send_keys("-!!!- ADD YOUR USERNAME HERE -!!!-")

    passwordInput = driver.find_element(By.XPATH, '//*[@id="loginform-password"]')
    passwordInput.send_keys("-!!!- ADD YOUR PASSWORD HERE -!!!-")

    submitLogin = driver.find_element(By.XPATH, '//*[@id="login-form"]/div[5]/button')
    submitLogin.click()

    time.sleep(5)

    i = 2 + int(offset)
    #loop through each SEETS entry in the data sheet
    while i < fileRows:
        if noDataVal.value > noDataLimit:
            break

        loop(lock, driver, i, noDataVal)
        i += 1
    driver.quit()


if __name__ == '__main__':
    #determines the amount of SEETS entries to search for
    ds = pandas.read_excel(inputSheet)
    fileRows = ds.shape[0] + 2

    lock = Lock()

    print("fileRows: " + str(fileRows))

    offset = ''
    while offset.isdigit() == False:
        offset = input("Offset Amount: (0 if you want to start from the beginning) ")

    #adds the labels to the spreadsheet for the new attributes
    book = openpyxl.load_workbook(inputSheet)
    sheet = book.active

    book.save(inputSheet)

    noDataVal = multiprocessing.Value('i', 0)

    process1 = Process(target=p1, args=(lock, fileRows, offset, noDataVal))

    process1.start()

    process1.join()