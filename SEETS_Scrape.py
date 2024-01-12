# Nichoals Wharton
# SEETS Scraping Program
# 12/17 updated version

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from datetime import datetime
import openpyxl
import time

now = datetime.now()
saveFile = 'SEETSDataSheet' + str(now.strftime("%m-%d-%H-%M")) + '.xlsx'

#create a webdriver instance and open up the page to scrape
element_list = []

# set the options to open the brower with 1920x1080 ar
options = Options()
options.add_argument("window-size=1920,1080")
driver = webdriver.Chrome(options=options)
driver.get("https://app.ecomsolutions.technology/seetsproducts/list")
print("Logging in...")
time.sleep(3)

book = openpyxl.Workbook()
sheet = book.active

popUp = driver.find_element(By.XPATH, '//*[@id="modalCookies"]/div/div/form/div[2]/button')
popUp.click()

time.sleep(1)

usernameInput = driver.find_element(By.XPATH, '//*[@id="Username"]')
usernameInput.send_keys("-!!!- ADD YOUR USERNAME HERE -!!!-")

passwordInput = driver.find_element(By.XPATH, '//*[@id="Password"]')
passwordInput.send_keys("-!!!- ADD YOUR PASSWORD HERE -!!!-")

submitLogin = driver.find_element(By.XPATH, '//*[@id="login-box-inner"]/form/div[5]/div/button')
submitLogin.click()

print("Login Succsessful!")
time.sleep(4)
input("Press Enter To Continue....")

maxItems = driver.find_element(By.ID, "paginationItemsPerPage")

itemDrop = Select(maxItems)
maxItemsVal = itemDrop.first_selected_option.get_attribute("value")

print("Max Items: " + maxItemsVal)

sheet.cell(row=1, column=1, value="JW Approval Color Code")
sheet.cell(row=1, column=2, value="MC Approval Color Code")
sheet.cell(row=1, column=3, value="ASIN")
sheet.cell(row=1, column=4, value="Hyper-ASIN")
sheet.cell(row=1, column=5, value="Amazon Title")
sheet.cell(row=1, column=6, value="Brand")
sheet.cell(row=1, column=7, value="SEETS Qty >10 Available Quantity")
sheet.cell(row=1, column=8, value="Keepa Sales >1 30 Day Sales")
sheet.cell(row=1, column=9, value="Unit Cost")
sheet.cell(row=1, column=10, value="Total Purchase Cost")
sheet.cell(row=1, column=11, value="Buy Box Price")
sheet.cell(row=1, column=12, value="Profit at BB")
sheet.cell(row=1, column=13, value="Remaining Quantity")

sheet.cell(row=1, column=14, value="Amazon Fee")
sheet.cell(row=1, column=15, value="FBA Fees")
sheet.cell(row=1, column=16, value="Proccessing Fee")
sheet.cell(row=1, column=17, value="Total Fees")
sheet.cell(row=1, column=18, value="Total Purchase Price (+) Total Fees")
sheet.cell(row=1, column=19, value="Buy Box Price (+) Total Fees")
sheet.cell(row=1, column=20, value="SAS Maximun Cost")
sheet.cell(row=1, column=21, value="Total Purchase Cost (-) SAS Maximun Cost")
sheet.cell(row=1, column=22, value="Lowest FBA")
sheet.cell(row=1, column=23, value="Lowest FBM")
sheet.cell(row=1, column=24, value="NET- Profit at BB")
sheet.cell(row=1, column=25, value="Average BB")
sheet.cell(row=1, column=26, value="Average Sales Price")
sheet.cell(row=1, column=27, value="Margin")
sheet.cell(row=1, column=28, value="ROI")
sheet.cell(row=1, column=29, value="MOQ")
sheet.cell(row=1, column=30, value="Total Sales")
sheet.cell(row=1, column=31, value="Stock Total")
sheet.cell(row=1, column=32, value="Total Sellers")
sheet.cell(row=1, column=33, value="FBA Sellers")
sheet.cell(row=1, column=34, value="Average Sales Rank")
sheet.cell(row=1, column=35, value="# Buyers / # Monthly Buyers")
sheet.cell(row=1, column=36, value="# Units / # Monthly Units")
sheet.cell(row=1, column=37, value="Date Added")
sheet.cell(row=1, column=38, value="Amazon Category")
sheet.cell(row=1, column=39, value="30 Day Sales")

sheet.cell(row=1, column=40, value="Average Sales Price")
sheet.cell(row=1, column=41, value="Average Lead Time")
sheet.cell(row=1, column=42, value="Buy Box Seller")
sheet.cell(row=1, column=43, value="Sales Rank")
sheet.cell(row=1, column=44, value="Second Buy Box Price")
sheet.cell(row=1, column=45, value="Profit At Second Buy Box")
sheet.cell(row=1, column=46, value="Multipack QTY")
sheet.cell(row=1, column=47, value="Reference Offer")
sheet.cell(row=1, column=48, value="Availability")
sheet.cell(row=1, column=49, value="Reviews")
sheet.cell(row=1, column=50, value="Source ID")
sheet.cell(row=1, column=51, value="Source Title")
sheet.cell(row=1, column=52, value="Merchant ID")



pf = driver.find_element(By.XPATH, '//*[@id="content-wrapper"]/div/div/div/div[2]/div/div/div/div')
print("products found: " + pf.text)
productsNum = ''
for i in pf.text:
    if i.isdigit():
        productsNum = productsNum + i

book.save(saveFile)
count = 0
currPage = 1
exitWhile = 0
while (1):
    book = openpyxl.load_workbook(saveFile)
    sheet = book.active
    
    for items in range(int(maxItemsVal)):
        print(str(count) + " of " + productsNum)
        
        try:
            driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[1]/table/tbody/tr[2]/td[2]')
        except NoSuchElementException:
            exitWhile = 1
            break


        sourceID = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[1]/table/tbody/tr[2]/td[2]')
        sourceTitle = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[1]/table/tbody/tr[3]/td[2]')
        amazonTitle = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[1]/table/tbody/tr[4]/td[2]')
        itemBrand = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[1]/table/tbody/tr[5]/td[2]')
        asin = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[1]/table/tbody/tr[6]/td[2]')
        reviews = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[1]/table/tbody/tr[7]/td[2]')
        availability = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[1]/table/tbody/tr[8]/td[2]')
        merchantID = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[1]/table/tbody/tr[9]/td[2]')
        dateAdded = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[1]/table/tbody/tr[10]/td[2]')
        moq = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[1]/table/tbody/tr[11]/td[2]')
        averageLeadTime = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[1]/table/tbody/tr[12]/td[2]')

        amazonCategory = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[2]/table/tbody/tr[1]/td[2]')
        amazonFee = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[2]/table/tbody/tr[2]/td[2]')
        totalSellers = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[2]/table/tbody/tr[3]/td[2]')
        fbaSellers = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[2]/table/tbody/tr[4]/td[2]')
        unitCost = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[2]/table/tbody/tr[5]/td[2]')
        multipackQty = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[2]/table/tbody/tr[6]/td[2]')
        totalPurchaseCost = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[2]/table/tbody/tr[7]/td[2]')
        fbaFees = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[2]/table/tbody/tr[8]/td[2]')
        buyBoxSeller = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[2]/table/tbody/tr[9]/td[2]')

        salesRank = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[3]/div/table/tbody/tr[1]/td[2]')
        averageSalesRank = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[3]/div/table/tbody/tr[2]/td[2]')
        buyBoxPrice = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[3]/div/table/tbody/tr[3]/td[2]')
        secondBuyBoxPrice = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[3]/div/table/tbody/tr[4]/td[2]')
        referenceOffer = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[3]/div/table/tbody/tr[5]/td[2]/input')
        profitAtBB = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[3]/div/table/tbody/tr[6]/td[2]')
        profitAt2ndBB = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[3]/div/table/tbody/tr[7]/td[2]')
        averageBB = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[3]/div/table/tbody/tr[8]/td[2]')
        roi = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[3]/div/table/tbody/tr[9]/td[2]')
        margin = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[3]/div/table/tbody/tr[10]/td[2]')
        buyersByMonthlyBuyers = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[3]/div/table/tbody/tr[11]/td[2]')
        unitsByMonthlyUnits = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[3]/div/table/tbody/tr[12]/td[2]')
        remQtyEle = driver.find_element(By.XPATH, '//*[@id="seetsItems"]/table/tbody/tr[' + str(int(items) + 1) + ']/td[1]/table/tbody/tr[8]/td[2]/a')
        
        remQtyStr = ''
        for ch in remQtyEle.get_attribute("data-original-title"):
            if ch.isdigit():
                remQtyStr += ch

        sheet.cell(row=(count + 2), column=3, value=asin.text) # Column C
        #4 hyper-asin
        sheet.cell(row=(count + 2), column=5, value=amazonTitle.text)
        sheet.cell(row=(count + 2), column=6, value=itemBrand.text)
        sheet.cell(row=(count + 2), column=7, value=("=IF(INT(M" + str(count) + ")>10,1,0)")) #seets quanity >10 Available Quantity
        sheet.cell(row=(count + 2), column=8, value=("=IF(INT(AM" + str(count) + ")>1,1,0)")) #keepa sales >1 30 Day Sales
        sheet.cell(row=(count + 2), column=9, value=unitCost.text)
        sheet.cell(row=(count + 2), column=10, value=totalPurchaseCost.text)
        sheet.cell(row=(count + 2), column=11, value=buyBoxPrice.text)
        sheet.cell(row=(count + 2), column=12, value=profitAtBB.text)
        sheet.cell(row=(count + 2), column=13, value=remQtyStr)

        sheet.cell(row=(count + 2), column=14, value=amazonFee.text)
        sheet.cell(row=(count + 2), column=15, value=fbaFees.text)
        sheet.cell(row=(count + 2), column=16, value="1.50") #processing fee
        sheet.cell(row=(count + 2), column=17, value=("=SUM(N" + str(count) + "+O" + str(count) + "+P" + str(count) + ")")) #total fees
        sheet.cell(row=(count + 2), column=18, value=("=SUM(J" + str(count) + "+Q" + str(count) + ")"))#18 Total Purchase Price (+) Total Fees
        sheet.cell(row=(count + 2), column=19, value=("=SUM(K" + str(count) + "+Q" + str(count) + ")"))#19 Buy Box Price (+) Total Fees
        #20 SAS Maximun Cost
        sheet.cell(row=(count + 2), column=21, value=("=SUM(J" + str(count) + "-T" + str(count) + ")"))#21 Total Purchase Cost (-) SAS Maximun Cost
        #22 Lowest FBA
        #23 Lowest FBM
        #24 NET- Profit at BB
        sheet.cell(row=(count + 2), column=25, value=averageBB.text)
        #26 average sales price
        sheet.cell(row=(count + 2), column=27, value=margin.text)
        sheet.cell(row=(count + 2), column=28, value=roi.text)
        sheet.cell(row=(count + 2), column=29, value=moq.text)
        #30 total sales
        #31 stock total
        sheet.cell(row=(count + 2), column=32, value=totalSellers.text)
        sheet.cell(row=(count + 2), column=33, value=fbaSellers.text)
        sheet.cell(row=(count + 2), column=34, value=averageSalesRank.text)
        sheet.cell(row=(count + 2), column=35, value=buyersByMonthlyBuyers.text)
        sheet.cell(row=(count + 2), column=36, value=unitsByMonthlyUnits.text)
        sheet.cell(row=(count + 2), column=37, value=dateAdded.text)
        sheet.cell(row=(count + 2), column=38, value=amazonCategory.text)
        #39 30 day sales
        #40 average sales price
        sheet.cell(row=(count + 2), column=41, value=averageLeadTime.text)
        sheet.cell(row=(count + 2), column=42, value=buyBoxSeller.text)
        sheet.cell(row=(count + 2), column=43, value=salesRank.text)
        sheet.cell(row=(count + 2), column=44, value=secondBuyBoxPrice.text)
        sheet.cell(row=(count + 2), column=45, value=profitAt2ndBB.text)
        sheet.cell(row=(count + 2), column=46, value=multipackQty.text)
        sheet.cell(row=(count + 2), column=47, value=referenceOffer.get_attribute('value'))
        sheet.cell(row=(count + 2), column=48, value=availability.text)
        sheet.cell(row=(count + 2), column=49, value=reviews.text)
        sheet.cell(row=(count + 2), column=50, value=sourceID.text)
        sheet.cell(row=(count + 2), column=51, value=sourceTitle.text)
        sheet.cell(row=(count + 2), column=52, value=merchantID.text)

        count += 1

    currPage += 1
    book.save(saveFile)

    if exitWhile == 1:
        break

    nextPage = driver.find_element(By.CSS_SELECTOR, '#paginationContainerDiv > div > ul > li.PagedList-skipToNext > a')
    if (nextPage.text == '»'):
        nextPage.click()
        time.sleep(10)
    else:
        break

    # Can be uncommented to ask the user if they want to continue after every 10 pages to stop the program without the chance of corrupting the data
    if currPage % 100 == 0:
        userIn = input('Do you want to continue?')
        
        if userIn == 'n':
            driver.quit()
            break